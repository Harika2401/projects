from flask import Flask, render_template, request, redirect, url_for, send_file
import cv2
import pytesseract
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

def preprocess_image(image_path):
    image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    _, binary_image = cv2.threshold(image, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    return binary_image

def perform_ocr(image):
    custom_config = r'--oem 3 --psm 6'
    text = pytesseract.image_to_string(image, config=custom_config)
    print("Extracted Text:")
    print(text)  # Debug: Print extracted text
    return text

def interpret_and_structure_text(text):
    lines = text.split('\n')
    headings = {}
    pattern = re.compile(r'^(.*?):\s*(.*)$')

    for line in lines:
        match = pattern.match(line.strip())
        if match:
            heading, value = match.groups()
            if heading not in headings:
                headings[heading] = []
            headings[heading].append(value)

    max_length = max(len(values) for values in headings.values())
    structured_data = []

    for i in range(max_length):
        row = []
        for heading in headings.keys():
            values = headings[heading]
            row.append(values[i] if i < len(values) else '')
        structured_data.append(row)

    print("Structured Data:")
    print(structured_data)  # Debug: Print structured data
    return list(headings.keys()), structured_data

def create_or_append_excel(headings, data, output_path):
    try:
        df_new = pd.DataFrame(data, columns=headings)
        if output_path.endswith('.xlsx'):
            if os.path.exists(output_path):
                # Load the existing workbook
                book = load_workbook(output_path)

                # Check if the sheet exists
                if 'Sheet1' in book.sheetnames:
                    # Load the existing worksheet
                    ws = book['Sheet1']

                    # Append the column headers if the worksheet is empty
                    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
                        for col_num, header in enumerate(headings, 1):
                            ws.cell(row=1, column=col_num, value=header)

                    # Find the next empty row in the worksheet
                    next_row = ws.max_row + 1

                    # Append new data to the worksheet
                    for row in dataframe_to_rows(df_new, index=False, header=False):
                        ws.append(row)

                    book.save(output_path)
                    print("Data appended successfully.")
                else:
                    # Create a new sheet if it does not exist
                    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                        df_new.to_excel(writer, index=False, sheet_name='Sheet1')
                    print("New sheet created and data written successfully.")
            else:
                # Create a new Excel file if it doesn't exist
                df_new.to_excel(output_path, index=False, sheet_name='Sheet1')
                print("New Excel file created with data.")
        else:
            raise ValueError("Output path must end with '.xlsx'")
    except ValueError as ve:
        print(f"Value error: {ve}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' in request.files:
        file = request.files['file']
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)
        process_image(file_path)
    elif 'camera' in request.form:
        capture_image_from_camera()

    return redirect(url_for('index'))

def capture_image_from_camera():
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        return "Error: Could not open camera."

    ret, frame = cap.read()
    if ret:
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], 'captured_image.jpg')
        cv2.imwrite(image_path, frame)
        cap.release()
        process_image(image_path)
    else:
        return "Error: Could not read frame."

def process_image(image_path):
    preprocessed_image = preprocess_image(image_path)
    extracted_text = perform_ocr(preprocessed_image)
    headings, structured_data = interpret_and_structure_text(extracted_text)
    output_path = 'output.xlsx'
    create_or_append_excel(headings, structured_data, output_path)

@app.route('/download')
def download_file():
    output_path = 'output.xlsx'
    return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)